"""Excel (.xlsx) export for YouTube market research results."""

from __future__ import annotations

from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .youtube_client import ChannelInfo, VideoInfo
from .keyword_extractor import KeywordEntry


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _write_header(ws, headers: list):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def _auto_width(ws, headers: list, col_widths: Optional[dict] = None):
    col_widths = col_widths or {}
    for i, h in enumerate(headers, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = col_widths.get(h, max(12, min(50, len(h) + 4)))


def _zebra(ws, n_rows: int, n_cols: int):
    for r in range(2, n_rows + 2):
        if r % 2 == 0:
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = ALT_FILL


def build_workbook(
    channel: ChannelInfo,
    channel_videos: list,
    keywords: list,
    top_by_keyword: dict,
    recent_by_keyword: dict,
    run_params: dict,
    output_path: str,
    tag_metrics: dict = None,
):
    wb = Workbook()

    # Tổng quan
    ws = wb.active
    ws.title = "Tổng quan"
    rows = [
        ("Tên kênh", channel.title),
        ("Handle", "@" + channel.handle if channel.handle and not channel.handle.startswith("@") else channel.handle),
        ("Channel ID", channel.channel_id),
        ("URL", channel.url),
        ("Người đăng ký", channel.subscriber_count),
        ("Tổng lượt xem", channel.view_count),
        ("Số video (kênh báo cáo)", channel.video_count),
        ("Số video đã lấy", len(channel_videos)),
        ("Từ khoá kênh (creator đặt)", ", ".join(channel.keywords) if channel.keywords else "(không có)"),
        ("Mô tả kênh", channel.description),
        ("", ""),
        ("--- Tham số chạy ---", ""),
        ("Thời điểm chạy", datetime.utcnow().isoformat() + "Z"),
        ("Backend (nguồn dữ liệu)", run_params.get("backend", "")),
        ("Khung thời gian (ngày)", run_params.get("days", "")),
        ("Số từ khoá", run_params.get("top_keywords", "")),
        ("Kết quả mỗi từ khoá", run_params.get("per_keyword", "")),
        ("Khu vực", run_params.get("region", "") or "(toàn cầu)"),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        c1 = ws.cell(row=i, column=1, value=k)
        c1.font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 100

    # Video của kênh
    ws = wb.create_sheet("Video của kênh")
    headers = ["Tiêu đề", "Ngày đăng", "Lượt xem", "Like", "Bình luận",
               "Thời lượng (s)", "Lượt xem/ngày", "Tags", "URL"]
    _write_header(ws, headers)
    sorted_videos = sorted(channel_videos, key=lambda v: v.published_at, reverse=True)
    for r, v in enumerate(sorted_videos, start=2):
        ws.cell(row=r, column=1, value=v.title)
        ws.cell(row=r, column=2, value=v.published_at[:10] if v.published_at else "")
        ws.cell(row=r, column=3, value=v.view_count)
        ws.cell(row=r, column=4, value=v.like_count)
        ws.cell(row=r, column=5, value=v.comment_count)
        ws.cell(row=r, column=6, value=v.duration_seconds)
        ws.cell(row=r, column=7, value=round(v.views_per_day, 1))
        ws.cell(row=r, column=8, value=", ".join(v.tags[:15]))
        ws.cell(row=r, column=9, value=v.url)
    _auto_width(ws, headers, {"Tiêu đề": 55, "Tags": 50, "URL": 45})
    _zebra(ws, len(sorted_videos), len(headers))

    # Từ khoá
    ws = wb.create_sheet("Từ khoá")
    headers = ["Hạng", "Từ khoá", "Điểm", "Nguồn", "Số video kênh chứa"]
    _write_header(ws, headers)
    for r, kw in enumerate(keywords, start=2):
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=kw.keyword)
        ws.cell(row=r, column=3, value=round(kw.score, 3))
        ws.cell(row=r, column=4, value=", ".join(sorted(kw.sources)))
        ws.cell(row=r, column=5, value=kw.channel_video_count)
    _auto_width(ws, headers, {"Từ khoá": 35, "Nguồn": 28})
    _zebra(ws, len(keywords), len(headers))

    # Top theo từ khoá
    ws = wb.create_sheet("Top theo từ khoá")
    headers = ["Từ khoá", "Hạng", "Tiêu đề", "Kênh", "Lượt xem", "Ngày đăng", "URL"]
    _write_header(ws, headers)
    row = 2
    for kw in keywords:
        vids = top_by_keyword.get(kw.keyword, [])
        for rank, v in enumerate(vids, start=1):
            ws.cell(row=row, column=1, value=kw.keyword)
            ws.cell(row=row, column=2, value=rank)
            ws.cell(row=row, column=3, value=v.title)
            ws.cell(row=row, column=4, value=v.channel_title)
            ws.cell(row=row, column=5, value=v.view_count)
            ws.cell(row=row, column=6, value=v.published_at[:10] if v.published_at else "")
            ws.cell(row=row, column=7, value=v.url)
            row += 1
    _auto_width(ws, headers, {"Tiêu đề": 55, "Kênh": 28, "URL": 45, "Từ khoá": 25})
    _zebra(ws, row - 2, len(headers))

    # Video mới nhiều view — kết quả chính
    ws = wb.create_sheet("Video mới nhiều view")
    headers = ["Từ khoá", "Hạng", "Tiêu đề", "Kênh", "Lượt xem", "Số ngày",
               "Lượt xem/ngày", "Ngày đăng", "URL"]
    _write_header(ws, headers)
    row = 2
    for kw in keywords:
        vids = recent_by_keyword.get(kw.keyword, [])
        # Sort theo LƯỢT XEM TUYỆT ĐỐI để thấy đối thủ nào có video mới hot nhất
        vids_sorted = sorted(vids, key=lambda v: v.view_count, reverse=True)
        for rank, v in enumerate(vids_sorted, start=1):
            ws.cell(row=row, column=1, value=kw.keyword)
            ws.cell(row=row, column=2, value=rank)
            ws.cell(row=row, column=3, value=v.title)
            ws.cell(row=row, column=4, value=v.channel_title)
            ws.cell(row=row, column=5, value=v.view_count)
            ws.cell(row=row, column=6, value=round(v.days_old, 1))
            ws.cell(row=row, column=7, value=round(v.views_per_day, 1))
            ws.cell(row=row, column=8, value=v.published_at[:10] if v.published_at else "")
            ws.cell(row=row, column=9, value=v.url)
            row += 1
    _auto_width(ws, headers, {"Tiêu đề": 55, "Kênh": 28, "URL": 45, "Từ khoá": 25})
    _zebra(ws, row - 2, len(headers))

    # Tag Analyzer sheet (nếu có data)
    if tag_metrics and tag_metrics.get("per_keyword"):
        ws = wb.create_sheet("Tag Analyzer")
        ta_headers = [
            "Từ khoá", "Competition", "Số kết quả YT",
            "Trend score (0-100)", "Xu hướng", "Sparkline 12 tháng",
            "Top 5 autosuggest",
        ]
        _write_header(ws, ta_headers)
        seo = tag_metrics.get("seo_summary") or {}
        per_kw = tag_metrics.get("per_keyword") or {}

        sorted_kws = sorted(per_kw.items(),
                            key=lambda x: -(x[1].get("trend", {}).get("score", 0)))
        for r, (kw, m) in enumerate(sorted_kws, start=2):
            comp = m.get("competition") or {}
            trend = m.get("trend") or {}
            sugs = m.get("autosuggest") or []
            ws.cell(row=r, column=1, value=kw)
            ws.cell(row=r, column=2, value=comp.get("level", "?").upper())
            ws.cell(row=r, column=3, value=comp.get("result_count", 0))
            ws.cell(row=r, column=4, value=trend.get("score", 0))
            ws.cell(row=r, column=5, value=trend.get("direction", "?"))
            ws.cell(row=r, column=6, value=trend.get("sparkline", "")[-40:])
            ws.cell(row=r, column=7, value=", ".join(sugs[:5]))
        _auto_width(ws, ta_headers, {"Từ khoá": 30, "Sparkline 12 tháng": 50,
                                       "Top 5 autosuggest": 60})
        _zebra(ws, len(sorted_kws), len(ta_headers))

        # Trang SEO Summary nếu có
        if seo and seo.get("video_count", 0) > 0:
            ws2 = wb.create_sheet("SEO Score chi tiết")
            seo_headers = ["Video Title", "SEO Score (/100)"]
            # Thêm cột cho mỗi component
            comp_keys = []
            if seo.get("videos") and seo["videos"]:
                comp_keys = list(seo["videos"][0]["components"].keys())
                for ck in comp_keys:
                    seo_headers.append(f"{ck} (điểm)")
                    seo_headers.append(f"{ck} (ghi chú)")
            _write_header(ws2, seo_headers)
            for r, v in enumerate(seo.get("videos", []), start=2):
                ws2.cell(row=r, column=1, value=v["video_title"])
                ws2.cell(row=r, column=2, value=v["score"])
                col = 3
                for ck in comp_keys:
                    c = v["components"].get(ck, {})
                    ws2.cell(row=r, column=col, value=c.get("score", 0))
                    ws2.cell(row=r, column=col + 1, value=c.get("note", ""))
                    col += 2
            _auto_width(ws2, seo_headers, {"Video Title": 50})

    # Sắp xếp thứ tự sheet: Tổng quan → Video mới nhiều view (kết quả chính)
    # → Top mọi thời → Tag Analyzer → SEO → Từ khoá → Video của kênh
    desired_order = [
        "Tổng quan", "Video mới nhiều view", "Top theo từ khoá",
        "Tag Analyzer", "SEO Score chi tiết",
        "Từ khoá", "Video của kênh",
    ]
    existing = {s.title: s for s in wb._sheets}
    wb._sheets = [existing[name] for name in desired_order if name in existing]

    wb.save(output_path)
    return output_path
